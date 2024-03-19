import openpyxl
from os import remove

file_name = 'manga_data.xlsx'

def open():
    try:
        wb = openpyxl.load_workbook(file_name)
        return wb
    except FileNotFoundError:
        print('No se encontró el archivo')
    except Exception as e:
        print(e)
    return None

def write(wb, data_dict):
    if not wb:
        print('Creando nuevo archivo...')
        wb = openpyxl.Workbook()
        ws = wb.active
    else:   
        ws = wb.create_sheet()

    for col, item in enumerate(data_dict.items(), 1):
        key, value = item
        ws.cell(row = 1, column = col, value = key)
        ws.cell(row = 2, column = col, value = value)

    return wb

def save(wb):
    try:
        wb.save(file_name)
        wb.close()
        print('Datos guardados con éxito')   

    except Exception as e:
        print('No se pudieron guardar los datos en el archivo')
        raise SystemExit(e)
    
def open_save(data_dict):
    try:
        wb = open()
        wb = write(wb, data_dict)
        save(wb)
    except Exception as e:
        print(e)

def read_sheet(ws):
    data = list(ws.iter_cols(values_only=True))
    print(f'{"~"*3 + " "*3}{data[0][1]}{" "*3 + "~"*3}')
    for idx, col in enumerate(data[1:], 1):
        print(f'{idx} - {col[0]}')
    return data

def del_sheet(wb, ws):
    if len(wb.get_sheet_names()) == 1:
        wb.close()
        remove(file_name)
        print('Se eliminaron todas las consultas')
        return True
    
    wb.remove_sheet(ws)
    return False